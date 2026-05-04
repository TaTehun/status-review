"""
Email body generation utilities.
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def format_xlsx(file_path: Path) -> None:
    """Apply auto-fit column widths and auto filter to an xlsx file."""
    wb = load_workbook(file_path)
    ws = wb.active

    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(file_path)


def build_tender_email_body(file_path: Path, date_str: str) -> str:
    """Build HTML email body for Tender Status report (date_str: mm/dd/yyyy)."""
    format_xlsx(file_path)

    df = pd.read_excel(file_path)

    cols = [c for c in ['Load ID', 'Tender Request Status'] if c in df.columns]
    df = df[cols]

    # Yellow bold headers
    header_style = 'background-color:#FFFF00; font-weight:bold; padding:4px 8px;'
    cell_style = 'padding:4px 8px;'

    headers = ''.join(f'<th style="{header_style}">{col}</th>' for col in df.columns)
    rows = ''
    for _, row in df.iterrows():
        cells = ''.join(f'<td style="{cell_style}">{val}</td>' for val in row)
        rows += f'<tr>{cells}</tr>'

    table_html = f'<table border="1" cellspacing="0" style="border-collapse:collapse;font-size:10pt;font-family:Calibri;"><tr>{headers}</tr>{rows}</table>'

    return (
        '<BODY style="font-size:11pt;font-family:Calibri">'
        f'Please see below for loads tendered on {date_str}.'
        '<br><br>'
        + table_html
        + '</BODY>'
    )
